#!/usr/bin/env python3
"""
build_ia_csv.py

Build the Hawk Academy IA Mapper deliverable from a JSON definition.

The script reads a single JSON file describing every URL in the inventory,
the vetting decisions for every keyword pulled from SE Ranking, and any
attendee-supplied metadata, then produces:

  - ia-map-{domain-slug}-{YYYY-MM-DD}.xlsx   (two sheets: IA Map + Vetting Log)
  - ia-map-{domain-slug}-{YYYY-MM-DD}.csv    (IA Map sheet as CSV mirror)
  - vetting-log-{domain-slug}-{YYYY-MM-DD}.csv (Vetting Log sheet as CSV mirror)

Usage:
    python3 build_ia_csv.py ia_data.json /path/to/output_dir/

JSON schema (keys marked * are required):

{
  "domain": "example.com.au",                          *
  "locale": "AU",                                        // defaults to "AU"
  "date": "2026-05-26",                                  // defaults to today
  "is_thin_site": false,                                 // true if Step 2b interview ran

  "urls": [                                             *
    {
      "url": "https://example.com/products/widget-a/",   *
      "section": "5. Product detail",                    // overrides URL-pattern default
      "page_title": "Widget A | Example",
      "is_admin": false,                                 // true = admin/policy page
      "is_proposed": false,                              // true = came from thin-site interview
      "notes": "200 staging products"
    }
  ],

  "keywords": [                                          *
    {
      "url": "https://example.com/products/widget-a/",   *
      "seed": "widget a",                                *
      "keyword": "widget a",                             *
      "msv": 1600,                                       // null/"TBC" if no data
      "cpc": 2.10,                                       // null/"TBC" if no data
      "intent": "commercial",                            // or "TBC"
      "kd": 30,                                          // null/"TBC" if no data
      "existing_position": 7,                            // null if not ranking
      "decision": "KEEP",                                * // "KEEP" or "REMOVED"
      "reason": "B2B match — keep"                       *
    }
  ]
}

If python-docx isn't installed... wait, this is xlsx. Install openpyxl:
    pip install openpyxl --break-system-packages

If openpyxl is unavailable the script falls back to CSV-only output so
the deliverable still ships.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit, urlparse


# ---------------------------------------------------------------------------
# Section classifier — URL path → (Section, Hierarchy depth)
# ---------------------------------------------------------------------------

SECTIONS = [
    # (label, predicate)
    ("1. Home",                lambda p: p in ("", "/")),
    ("2. Sectors",             lambda p: p.startswith("/sectors/") or p.startswith("/industries/")),
    ("3. Brands",              lambda p: p.startswith("/brands/") or p.startswith("/partners/")),
    ("4. Categories / Services", lambda p: (p.startswith("/collections/")
                                              or p.startswith("/product-categories/")
                                              or p.startswith("/services/")
                                              or p.startswith("/category/")
                                              or p.startswith("/categories/"))),
    ("5. Product detail",      lambda p: p.startswith("/products/") or p.startswith("/product/")),
    ("6. Locations",           lambda p: (p.startswith("/locations/")
                                            or p.startswith("/cities/")
                                            or p.startswith("/areas/"))),
    ("7. Blog / Resources",    lambda p: (p.startswith("/blog/")
                                            or p.startswith("/blogs/")
                                            or p.startswith("/articles/")
                                            or p.startswith("/insights/")
                                            or p.startswith("/news/")
                                            or p.startswith("/resources/")
                                            or p.startswith("/guides/")
                                            or p.startswith("/learn/"))),
    ("8. Static pages",        lambda p: (p.startswith("/about")
                                            or p.startswith("/team")
                                            or p.startswith("/contact")
                                            or p.startswith("/careers")
                                            or p.startswith("/jobs"))),
    ("9. Admin / Policy",      lambda p: any(s in p for s in (
        "/privacy", "/terms", "/policies", "/policy", "/shipping", "/returns",
        "/refund", "/cookie", "/disclaimer", "/accessibility",
        "/cart", "/checkout", "/login", "/register", "/account", "/sign-in",
        "/sign-up", "/wishlist", "/my-account",
        "/404", "/search", "/thank-you", "/order-confirmation", "/sitemap",
    ))),
]

ADMIN_SECTION_LABEL = "9. Admin / Policy"

SECTION_ORDER = [label for label, _ in SECTIONS] + ["10. Other"]


def classify_section(url: str) -> str:
    path = urlparse(url).path.lower().rstrip("/")
    if path == "":
        path = "/"
    for label, predicate in SECTIONS:
        if predicate(path):
            return label
    return "10. Other"


def hierarchy_depth(url: str) -> int:
    """0 for home, 1 for top-level (e.g. /about/), 2 for /products/, etc."""
    path = urlparse(url).path.strip("/")
    if not path:
        return 0
    return len(path.split("/"))


def normalise_url(url: str) -> str:
    """Canonicalise: lowercase host, single trailing slash on directory paths."""
    if not url:
        return url
    s = urlsplit(url.strip())
    path = s.path
    if path.endswith("/") and len(path) > 1:
        path = path  # already canonical
    elif not path.endswith("/") and "." not in path.rsplit("/", 1)[-1]:
        path = path + "/"
    return urlunsplit((s.scheme, s.netloc.lower(), path, s.query, ""))


def domain_slug(domain: str) -> str:
    """example.com.au -> example-com-au"""
    return re.sub(r"[^a-z0-9]+", "-", domain.lower()).strip("-")


def parse_int(v) -> int:
    """SE Ranking sometimes serialises MSV as a string with commas or 'TBC'."""
    if v in (None, "", "TBC"):
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().lower().replace(",", "")
    if s.endswith("k"):
        try:
            return int(float(s[:-1]) * 1000)
        except ValueError:
            return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


# ---------------------------------------------------------------------------
# Build phase
# ---------------------------------------------------------------------------

IA_HEADERS = [
    "Section", "Hierarchy depth", "URL", "Page title", "Status",
    "Primary keyword", "Primary MSV", "Existing position",
    "Supporting keywords (top 7)", "Cluster volume", "Intent", "Notes",
]

VETTING_HEADERS = [
    "URL", "Seed", "Keyword", "MSV", "Intent", "Decision", "Reason",
]


def build_ia_rows(data: dict) -> list:
    """Return one row per URL for the IA Map sheet."""
    # Group KEEP keywords by URL
    keep_by_url = defaultdict(list)
    for kw in data.get("keywords", []):
        if kw.get("decision", "").upper() != "KEEP":
            continue
        url = normalise_url(kw.get("url", ""))
        if not url:
            continue
        keep_by_url[url].append(kw)

    rows = []
    for url_meta in data.get("urls", []):
        url = normalise_url(url_meta.get("url", ""))
        if not url:
            continue

        section = url_meta.get("section") or classify_section(url)
        is_admin = bool(url_meta.get("is_admin")) or section == ADMIN_SECTION_LABEL
        is_proposed = bool(url_meta.get("is_proposed"))
        kws = sorted(keep_by_url.get(url, []), key=lambda k: -parse_int(k.get("msv")))

        if is_admin:
            status = "Not mapped — admin/policy page"
            primary_kw = ""
            primary_msv = ""
            existing_pos = ""
            supporting = ""
            cluster_vol = 0
            intent = ""
        elif not kws:
            status = "Proposed (thin-site IA)" if is_proposed else "Gap — needs content"
            primary_kw = ""
            primary_msv = ""
            existing_pos = ""
            supporting = ""
            cluster_vol = 0
            intent = ""
        else:
            primary = kws[0]
            primary_kw = primary.get("keyword", "")
            primary_msv = parse_int(primary.get("msv"))
            existing_pos = primary.get("existing_position")
            if existing_pos in (None, ""):
                existing_pos = ""
            supporting = "\n".join(
                "{0} ({1})".format(k.get("keyword", ""), parse_int(k.get("msv")))
                for k in kws[1:8]
            )
            cluster_vol = sum(parse_int(k.get("msv")) for k in kws)
            intent = primary.get("intent", "") or ""
            status = "Proposed (thin-site IA)" if is_proposed else "Existing"

        rows.append({
            "Section": section,
            "Hierarchy depth": hierarchy_depth(url),
            "URL": url,
            "Page title": url_meta.get("page_title", ""),
            "Status": status,
            "Primary keyword": primary_kw,
            "Primary MSV": primary_msv,
            "Existing position": existing_pos,
            "Supporting keywords (top 7)": supporting,
            "Cluster volume": cluster_vol,
            "Intent": intent,
            "Notes": url_meta.get("notes", ""),
        })

    # Hierarchical sort: Section index -> URL alpha -> Cluster vol desc
    section_idx = {s: i for i, s in enumerate(SECTION_ORDER)}

    def sort_key(r):
        return (
            section_idx.get(r["Section"], 99),
            r["URL"],
            -int(r["Cluster volume"] or 0),
        )

    rows.sort(key=sort_key)
    return rows


def build_vetting_rows(data: dict) -> list:
    """Return one row per keyword for the Vetting Log sheet."""
    rows = []
    for kw in data.get("keywords", []):
        rows.append({
            "URL": normalise_url(kw.get("url", "")),
            "Seed": kw.get("seed", ""),
            "Keyword": kw.get("keyword", ""),
            "MSV": parse_int(kw.get("msv")) if kw.get("msv") not in (None, "TBC") else "TBC",
            "Intent": kw.get("intent", "") or "",
            "Decision": (kw.get("decision") or "").upper(),
            "Reason": kw.get("reason", ""),
        })
    # KEEP first, then by descending MSV
    def vsort(r):
        msv = r["MSV"] if isinstance(r["MSV"], int) else 0
        return (r["Decision"] != "KEEP", -msv)
    rows.sort(key=vsort)
    return rows


# ---------------------------------------------------------------------------
# Output phase
# ---------------------------------------------------------------------------

def write_csv(path: Path, headers: list, rows: list) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(path: Path, ia_rows: list, vetting_rows: list) -> bool:
    """Return True on success, False if openpyxl isn't available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()

    # --- IA Map sheet ---
    ws = wb.active
    ws.title = "IA Map"
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(IA_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for r in ia_rows:
        ws.append([r[h] for h in IA_HEADERS])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap

    widths = [22, 8, 50, 32, 28, 28, 12, 12, 50, 12, 16, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # --- Vetting Log sheet ---
    vws = wb.create_sheet(title="Vetting Log")
    vws.append(VETTING_HEADERS)
    for cell in vws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for r in vetting_rows:
        vws.append([r[h] for h in VETTING_HEADERS])
    for row in vws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap

    vwidths = [50, 24, 32, 10, 16, 12, 40]
    for i, w in enumerate(vwidths, 1):
        vws.column_dimensions[get_column_letter(i)].width = w
    vws.freeze_panes = "A2"

    wb.save(path)
    return True


def main() -> int:
    ap = argparse.ArgumentParser(description="Build the Hawk Academy IA Mapper deliverable.")
    ap.add_argument("data_json", help="Path to ia_data.json")
    ap.add_argument("out_dir", help="Output directory (a workspace folder)")
    args = ap.parse_args()

    data_path = Path(args.data_json)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(data_path, encoding="utf-8") as f:
        data = json.load(f)

    domain = data.get("domain") or "site"
    date_str = data.get("date") or date.today().isoformat()
    slug = "{0}-{1}".format(domain_slug(domain), date_str)

    ia_rows = build_ia_rows(data)
    vetting_rows = build_vetting_rows(data)

    xlsx_path = out_dir / "ia-map-{0}.xlsx".format(slug)
    csv_ia_path = out_dir / "ia-map-{0}.csv".format(slug)
    csv_vet_path = out_dir / "vetting-log-{0}.csv".format(slug)

    write_csv(csv_ia_path, IA_HEADERS, ia_rows)
    write_csv(csv_vet_path, VETTING_HEADERS, vetting_rows)

    xlsx_ok = write_xlsx(xlsx_path, ia_rows, vetting_rows)

    print("# IA rows:       {0}".format(len(ia_rows)), file=sys.stderr)
    print("# Vetting rows:  {0}".format(len(vetting_rows)), file=sys.stderr)
    print("# CSV (IA):      {0}".format(csv_ia_path), file=sys.stderr)
    print("# CSV (vetting): {0}".format(csv_vet_path), file=sys.stderr)
    if xlsx_ok:
        print("# XLSX:          {0}".format(xlsx_path), file=sys.stderr)
    else:
        print("# XLSX skipped — openpyxl not installed. Install with:", file=sys.stderr)
        print("#   pip install openpyxl --break-system-packages", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
